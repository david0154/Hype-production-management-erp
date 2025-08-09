from tkinter import *
from tkinter import filedialog, messagebox
from tkinter.ttk import Treeview, OptionMenu
import sqlite3
import os
from datetime import date, datetime
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

# Create safe writable path
APP_FOLDER = os.path.join(os.getenv('APPDATA'), 'HypeProduction')
os.makedirs(APP_FOLDER, exist_ok=True)

DB_NAME = os.path.join(APP_FOLDER, 'database.db')

def init_db():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS entries (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        article TEXT, card TEXT, color TEXT, size TEXT,
        qty INTEGER, component TEXT, print_opt TEXT, date TEXT
    )''')
    conn.commit()
    conn.close()

def show_main_ui():
    root = Tk()
    root.title("Dashboard - Hype Production Management")
    root.geometry("1000x750")
    root.configure(bg="#e0f7fa")

    # --- Define functions that will be used by widgets ---

    def update_dashboard(filtered_data=None):
        # Clear existing treeview items
        for row_in_tree in tree.get_children():
            tree.delete(row_in_tree)

        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()

        query = "SELECT * FROM entries WHERE 1=1"
        values = []

        if filtered_data:
            if filtered_data.get("article"):
                query += " AND article LIKE ?"
                values.append(f"%{filtered_data['article']}%")
            if filtered_data.get("card"):
                query += " AND card LIKE ?"
                values.append(f"%{filtered_data['card']}%")
            if filtered_data.get("print_opt"):
                query += " AND print_opt = ?"
                values.append(filtered_data["print_opt"])

            start_date_val = filtered_data.get("start_date")
            end_date_val = filtered_data.get("end_date")

            if start_date_val:
                try:
                    datetime.strptime(start_date_val, '%Y-%m-%d')
                    query += " AND date >= ?"
                    values.append(start_date_val)
                except ValueError:
                    messagebox.showwarning("Filter Error", f"Invalid start date format: {start_date_val}. Use<ctrl97>MM-DD.")
            if end_date_val:
                try:
                    datetime.strptime(end_date_val, '%Y-%m-%d')
                    query += " AND date <= ?"
                    values.append(end_date_val)
                except ValueError:
                    messagebox.showwarning("Filter Error", f"Invalid end date format: {end_date_val}. Use<ctrl97>MM-DD.")

        c.execute(query, values)
        rows = c.fetchall()
        for r_data in rows:
            tree.insert('', END, values=r_data)
        conn.close()

    def save_entry():
        data = (
            article_entry.get(), card_entry.get(), color_entry.get(),
            size_entry.get(), qty_entry.get(), component_entry.get(),
            print_var.get(), str(date.today())
        )
        if not article_entry.get():
            messagebox.showerror("Error", "Article fields cannot be empty.")
            return

        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute("INSERT INTO entries (article, card, color, size, qty, component, print_opt, date) VALUES (?, ?, ?, ?, ?, ?, ?, ?)", data)
        conn.commit()
        conn.close()
        messagebox.showinfo("Saved", "Entry saved successfully.")
        article_entry.delete(0, END)
        card_entry.delete(0, END)
        color_entry.delete(0, END)
        size_entry.delete(0, END)
        qty_entry.delete(0, END)
        component_entry.delete(0, END)
        print_var.set("Yes")
        update_dashboard()

    def upload_image():
        filepath = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg *.png *.jpeg")])
        if not filepath: return
        article = article_entry.get()
        if not article:
            messagebox.showerror("Error", "Enter article number first to name the image.")
            return

        safe_article_name = "".join(c if c.isalnum() else "_" for c in article)
        os.makedirs("images", exist_ok=True)
        original_extension = filepath.split('.')[-1]
        filename = f"images/{safe_article_name}.{original_extension}"

        try:
            with open(filepath, "rb") as src, open(filename, "wb") as dst:
                dst.write(src.read())
            messagebox.showinfo("Uploaded", f"Image uploaded as {filename}")
        except Exception as e:
            messagebox.showerror("Upload Error", f"Failed to upload image: {e}")

    def search_entries():
        start_date_val = search_start_date_entry.get()
        end_date_val = search_end_date_entry.get()

        if start_date_val:
            try:
                datetime.strptime(start_date_val, '%Y-%m-%d')
            except ValueError:
                messagebox.showerror("Date Error", "Invalid Start Date format. Please use<ctrl97>MM-DD.")
                return
        if end_date_val:
            try:
                datetime.strptime(end_date_val, '%Y-%m-%d')
            except ValueError:
                messagebox.showerror("Date Error", "Invalid End Date format. Please use<ctrl97>MM-DD.")
                return

        filters = {
            "article": search_article_entry.get(),
            "card": search_card_entry.get(),
            "print_opt": search_print_var.get() if search_print_var.get() != "All" else "",
            "start_date": start_date_val,
            "end_date": end_date_val
        }
        update_dashboard(filters)

    def edit_entry():
        selected_items = tree.selection()
        if not selected_items:
            messagebox.showerror("Error", "Select an entry to edit.")
            return

        selected_item_id_tree = selected_items[0]
        item_values = tree.item(selected_item_id_tree)["values"]

        edit_window = Toplevel(root)
        edit_window.title("Edit Entry")
        edit_window.geometry("350x380")
        edit_window.configure(bg="#e0f7fa")
        edit_window.grab_set()

        def update_db_entry():
            new_date_val = edit_date_var.get()
            try:
                if new_date_val:
                    datetime.strptime(new_date_val, '%Y-%m-%d')
            except ValueError:
                messagebox.showerror("Date Error", "Invalid date format for entry. Please use<ctrl97>MM-DD.", parent=edit_window)
                return

            updated_data = (
                edit_article_var.get(), edit_card_var.get(), edit_color_var.get(),
                edit_size_var.get(), edit_qty_var.get(), edit_component_var.get(),
                edit_print_opt_var.get(), new_date_val,
                item_values[0]
            )
            conn = sqlite3.connect(DB_NAME)
            c = conn.cursor()
            c.execute("""UPDATE entries SET article=?, card=?, color=?, size=?, qty=?, component=?, print_opt=?, date=?
                         WHERE id=?""", updated_data)
            conn.commit()
            conn.close()
            messagebox.showinfo("Updated", "Entry updated successfully.", parent=edit_window)
            edit_window.destroy()
            update_dashboard()

        labels_text = ["Article:", "Card:", "Color:", "Size:", "Qty:", "Component:", "Print:", "Date (YYYY-MM-DD):"]

        edit_article_var = StringVar(value=item_values[1])
        edit_card_var = StringVar(value=item_values[2])
        edit_color_var = StringVar(value=item_values[3])
        edit_size_var = StringVar(value=item_values[4])
        edit_qty_var = StringVar(value=item_values[5])
        edit_component_var = StringVar(value=item_values[6])
        edit_print_opt_var = StringVar(value=item_values[7])
        edit_date_var = StringVar(value=item_values[8])

        entry_vars_map = {
            "Article:": edit_article_var, "Card:": edit_card_var, "Color:": edit_color_var,
            "Size:": edit_size_var, "Qty:": edit_qty_var, "Component:": edit_component_var,
        }

        edit_entries_frame = Frame(edit_window, bg="#e0f7fa")
        edit_entries_frame.pack(pady=10, padx=10, fill="x", expand=True)

        for i, text in enumerate(labels_text):
            Label(edit_entries_frame, text=text, bg="#e0f7fa").grid(row=i, column=0, sticky="w", pady=3, padx=5)
            if text == "Print:":
                OptionMenu(edit_entries_frame, edit_print_opt_var, item_values[7], "Yes", "No").grid(row=i, column=1, sticky="ew", pady=3, padx=5)
            elif text == "Date (YYYY-MM-DD):":
                 Entry(edit_entries_frame, textvariable=edit_date_var).grid(row=i, column=1, sticky="ew", pady=3, padx=5)
            else:
                Entry(edit_entries_frame, textvariable=entry_vars_map[text]).grid(row=i, column=1, sticky="ew", pady=3, padx=5)

        edit_entries_frame.columnconfigure(1, weight=1)
        Button(edit_window, text="Update", command=update_db_entry, bg="#00796b", fg="white").pack(pady=10)

    def delete_entry():
        selected_items = tree.selection()
        if not selected_items:
            messagebox.showerror("Error", "Select an entry to delete.")
            return

        selected_item_id_tree = selected_items[0]
        item_id_db = tree.item(selected_item_id_tree)["values"][0]

        if messagebox.askyesno("Confirm Delete", "Are you sure you want to delete this entry?"):
            conn = sqlite3.connect(DB_NAME)
            c = conn.cursor()
            c.execute("DELETE FROM entries WHERE id=?", (item_id_db,))
            conn.commit()
            conn.close()
            messagebox.showinfo("Deleted", "Entry deleted successfully.")
            update_dashboard()

    def get_current_filters_for_export():
        start_val = search_start_date_entry.get()
        end_val = search_end_date_entry.get()
        valid_start_date = None
        valid_end_date = None

        if start_val:
            try:
                datetime.strptime(start_val, '%Y-%m-%d')
                valid_start_date = start_val
            except ValueError: pass
        if end_val:
            try:
                datetime.strptime(end_val, '%Y-%m-%d')
                valid_end_date = end_val
            except ValueError: pass

        return {
            "article": search_article_entry.get(),
            "card": search_card_entry.get(),
            "print_opt": search_print_var.get() if search_print_var.get() != "All" else "",
            "start_date": valid_start_date,
            "end_date": valid_end_date
        }

    def export_excel():
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if not filepath: return

        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()

        query = "SELECT id, article, card, color, size, qty, component, print_opt, date FROM entries WHERE 1=1"
        values = []
        current_filters = get_current_filters_for_export()

        if current_filters.get("article"):
            query += " AND article LIKE ?"
            values.append(f"%{current_filters['article']}%")
        if current_filters.get("card"):
            query += " AND card LIKE ?"
            values.append(f"%{current_filters['card']}%")
        if current_filters.get("print_opt"):
            query += " AND print_opt = ?"
            values.append(current_filters["print_opt"])
        if current_filters.get("start_date"):
            query += " AND date >= ?"
            values.append(current_filters["start_date"])
        if current_filters.get("end_date"):
            query += " AND date <= ?"
            values.append(current_filters["end_date"])

        c.execute(query, values)
        rows = c.fetchall()
        conn.close()

        wb = Workbook()
        ws = wb.active
        ws.append(["Article", "Card", "Color", "Size", "Qty", "Component", "Print", "Date"])
        for row in rows:
            ws.append(row)

        try:
            wb.save(filepath)
            messagebox.showinfo("Exported", "Data exported to Excel successfully.")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export to Excel: {e}")

    def export_pdf():
        filepath = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        if not filepath: return

        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        query = "SELECT ID, article, card, color, size, qty, component, print_opt, date FROM entries WHERE 1=1"
        values = []
        current_filters = get_current_filters_for_export()

        if current_filters.get("article"):
            query += " AND article LIKE ?"
            values.append(f"%{current_filters['article']}%")
        if current_filters.get("card"):
            query += " AND card LIKE ?"
            values.append(f"%{current_filters['card']}%")
        if current_filters.get("print_opt"):
            query += " AND print_opt = ?"
            values.append(current_filters["print_opt"])
        if current_filters.get("start_date"):
            query += " AND date >= ?"
            values.append(current_filters["start_date"])
        if current_filters.get("end_date"):
            query += " AND date <= ?"
            values.append(current_filters["end_date"])

        c.execute(query, values)
        db_rows = c.fetchall()
        conn.close()

        pdf_canvas = canvas.Canvas(filepath, pagesize=A4)
        width, height = A4
        margin = 40
        x_start_pos = margin
        y_pos = height - margin
        line_height = 16
        header_line_height = 20

        pdf_canvas.setFont("Helvetica-Bold", 14)
        pdf_canvas.drawString(x_start_pos, y_pos, "Laser Entries")
        y_pos -= (header_line_height * 1.5)

        headers = ["Article", "Card", "Color", "Size", "Qty", "Component", "Print", "Date"]

        available_width = width - 2 * margin
        col_widths_dict = { "ID": 0.06, "Article": 0.18, "Card": 0.12, "Color": 0.10,
                            "Size": 0.08, "Qty": 0.07, "Component": 0.18, "Print": 0.09, "Date": 0.12 }
        actual_col_widths = [available_width * col_widths_dict.get(h, 0.1) for h in headers]

        def draw_headers_on_page(canvas_obj, current_y_pos):
            canvas_obj.setFont("Helvetica-Bold", 9)
            current_x_pos = x_start_pos
            for i_h, header_text in enumerate(headers):
                canvas_obj.drawString(current_x_pos, current_y_pos, header_text)
                current_x_pos += actual_col_widths[i_h]
            current_y_pos -= (line_height * 0.5)
            canvas_obj.line(x_start_pos, current_y_pos, width - margin, current_y_pos)
            current_y_pos -= (line_height * 0.8)
            return current_y_pos

        y_pos = draw_headers_on_page(pdf_canvas, y_pos)
        pdf_canvas.setFont("Helvetica", 8)

        for row_data in db_rows:
            if y_pos < margin + line_height:
                pdf_canvas.showPage()
                pdf_canvas.setFont("Helvetica-Bold", 14)
                pdf_canvas.drawString(x_start_pos, height - margin, "LASER Entries (Continued)")
                y_pos = height - margin - (header_line_height * 1.5)
                y_pos = draw_headers_on_page(pdf_canvas, y_pos)
                pdf_canvas.setFont("Helvetica", 8)

            current_x_pos = x_start_pos
            for i_val, value in enumerate(row_data):
                cell_text = str(value)
                max_chars = int(actual_col_widths[i_val] / (pdf_canvas._fontsize * 0.5))
                if len(cell_text) > max_chars and max_chars > 3:
                     cell_text = cell_text[:max_chars-3] + "..."
                pdf_canvas.drawString(current_x_pos, y_pos, cell_text)
                current_x_pos += actual_col_widths[i_val]
            y_pos -= line_height

        try:
            pdf_canvas.save()
            messagebox.showinfo("Exported", "Data exported to PDF successfully.")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export to PDF: {e}")

    def show_import_mapping_window(filepath, excel_headers):
        mapping_window = Toplevel(root)
        mapping_window.title("Map Excel Columns")
        mapping_window.geometry("400x400")
        mapping_window.configure(bg="#e0f7fa")
        mapping_window.grab_set()

        mapping_frame = Frame(mapping_window, bg="#e0f7fa")
        mapping_frame.pack(pady=10, padx=10, fill="both", expand=True)

        db_fields = {
            "Article": True,
            "Card": False, # Changed Card to be NOT required
            "Color": False,
            "Size": False,
            "Qty": False,
            "Component": False,
            "Print": False,
            "Date": False
        }

        mapping_vars = {}

        Label(mapping_frame, text="Select Excel column for each field:", font=("Arial", 10, "bold"), bg="#e0f7fa").grid(row=0, column=0, columnspan=2, pady=10)

        row_num = 1
        excel_header_options = ["-- Skip --"] + excel_headers

        for field_name, required in db_fields.items():
            label_text = f"{field_name}:"
            if required:
                label_text += " (Required)"

            Label(mapping_frame, text=label_text, bg="#e0f7fa").grid(row=row_num, column=0, sticky="w", padx=5, pady=3)
            mapping_vars[field_name] = StringVar(value=excel_header_options[0]) # Default to -- Skip --
            
            # Attempt to pre-select based on likely header names (case-insensitive)
            for header in excel_headers:
                 if header and header.lower() == field_name.lower():
                      mapping_vars[field_name].set(header)
                      break

            # Specific default guesses for Qty, Print, Date if exact match not found
            if field_name == "Qty":
                 for header in excel_headers:
                     if header and "quant" in header.lower(): # Matches "QUANTITY"
                          mapping_vars[field_name].set(header)
                          break
            if field_name == "Print":
                 for header in excel_headers:
                     if header and ("print" in header.lower()):
                          mapping_vars[field_name].set(header)
                          break
            if field_name == "Date":
                 for header in excel_headers:
                     if header and ("date" in header.lower()):
                          mapping_vars[field_name].set(header)
                          break


            OptionMenu(mapping_frame, mapping_vars[field_name], *excel_header_options).grid(row=row_num, column=1, sticky="ew", padx=5, pady=3)
            row_num += 1

        mapping_frame.columnconfigure(1, weight=1)

        def start_import():
            selected_mapping = {field: var.get() for field, var in mapping_vars.items()}

            # Validate ONLY truly required fields are mapped
            for field_name, required in db_fields.items():
                if required and selected_mapping[field_name] == "-- Skip --":
                    messagebox.showerror("Mapping Error", f"'{field_name}' field is required but not mapped.", parent=mapping_window)
                    return

            mapping_window.destroy()
            perform_import(filepath, selected_mapping, excel_headers)


        Button(mapping_window, text="Import Data", command=start_import, bg="#00796b", fg="white").pack(pady=10)


    def perform_import(filepath, mapping, excel_headers):
        try:
            workbook = load_workbook(filepath)
            sheet = workbook.active

            # Excel headers might be in row 2, data starts row 3.
            # Need to find the actual start row for data if headers are not row 1.
            # Assuming headers are in row 2 based on your image.
            data_start_row = 3
            # We already got headers from row 2 in import_excel()

            header_to_index = {header: index for index, header in enumerate(excel_headers)}
            column_indices = {}
            for db_field, excel_header in mapping.items():
                if excel_header != "-- Skip --" and excel_header in header_to_index:
                    column_indices[db_field] = header_to_index[excel_header]
                else:
                    column_indices[db_field] = -1 # Indicates this field should be skipped or defaulted


            conn = sqlite3.connect(DB_NAME)
            c = conn.cursor()
            imported_count = 0
            errors = []

            # Iterate through rows starting from the determined data_start_row
            for row_index in range(data_start_row, sheet.max_row + 1):
                row_values = [sheet.cell(row=row_index, column=col+1).value for col in range(sheet.max_column)]

                try:
                    # Get data using mapped column indices
                    # Use .strip() and default to "" for text fields if column is skipped or empty
                    article = str(row_values[column_indices["Article"]] if column_indices["Article"] != -1 and column_indices["Article"] < len(row_values) and row_values[column_indices["Article"]] is not None else "").strip()
                    card = str(row_values[column_indices["Card"]] if column_indices["Card"] != -1 and column_indices["Card"] < len(row_values) and row_values[column_indices["Card"]] is not None else "").strip() # Card is now optional
                    color = str(row_values[column_indices["Color"]] if column_indices["Color"] != -1 and column_indices["Color"] < len(row_values) and row_values[column_indices["Color"]] is not None else "").strip()
                    size = str(row_values[column_indices["Size"]] if column_indices["Size"] != -1 and column_indices["Size"] < len(row_values) and row_values[column_indices["Size"]] is not None else "").strip()
                    qty = str(row_values[column_indices["Qty"]] if column_indices["Qty"] != -1 and column_indices["Qty"] < len(row_values) and row_values[column_indices["Qty"]] is not None else "").strip()
                    component = str(row_values[column_indices["Component"]] if column_indices["Component"] != -1 and column_indices["Component"] < len(row_values) and row_values[column_indices["Component"]] is not None else "").strip()


                    # Handle Print field mapping and validation
                    print_opt = "No" # Default if skipped or invalid
                    if column_indices["Print"] != -1 and column_indices["Print"] < len(row_values):
                         print_opt_raw = row_values[column_indices["Print"]]
                         if print_opt_raw is not None:
                             print_opt = str(print_opt_raw).strip().capitalize()
                             if print_opt not in ["Yes", "No"]:
                                  errors.append(f"Row {row_index}: Invalid 'Print' value '{print_opt_raw}'. Using 'No'.")
                                  print_opt = "No" # Default if value is something else


                    # Handle Date field mapping and validation
                    entry_date = str(date.today()) # Default if skipped or parsing fails
                    if column_indices["Date"] != -1 and column_indices["Date"] < len(row_values):
                         date_raw = row_values[column_indices["Date"]]
                         if date_raw is not None:
                             if isinstance(date_raw, datetime):
                                 entry_date = date_raw.strftime('%Y-%m-%d')
                             elif isinstance(date_raw, date):
                                  entry_date = date_raw.strftime('%Y-%m-%d')
                             elif isinstance(date_raw, (int, float)): # Handle dates stored as numbers by Excel
                                  try:
                                       # Convert Excel serial date number to datetime
                                       # Need to handle 1900 vs 1904 date system if necessary
                                       from openpyxl.utils.datetime import from_excel
                                       entry_date_obj = from_excel(date_raw)
                                       entry_date = entry_date_obj.strftime('%Y-%m-%d')
                                  except ValueError:
                                       errors.append(f"Row {row_index}: Cannot convert numeric date '{date_raw}'. Using today's date.")
                                       entry_date = str(date.today())

                             elif isinstance(date_raw, str) and date_raw:
                                 try:
                                     # Attempt to parse various string formats
                                     # Handles ISO format (e.g., "2023-10-27T10:00:00Z")
                                     if 'T' in date_raw and ('+' in date_raw or 'Z' in date_raw):
                                          entry_date = datetime.fromisoformat(date_raw.replace("Z", "+00:00")).strftime('%Y-%m-%d')
                                     else:
                                         # Attempt standard YYYY-MM-DD first, then common variations
                                         try:
                                              entry_date = datetime.strptime(date_raw, '%Y-%m-%d').strftime('%Y-%m-%d')
                                         except ValueError:
                                             try:
                                                 entry_date = datetime.strptime(date_raw, '%m/%d/%Y').strftime('%Y-%m-%d')
                                             except ValueError:
                                                 try:
                                                     entry_date = datetime.strptime(date_raw, '%d/%m/%Y').strftime('%Y-%m-%d')
                                                 except ValueError:
                                                     try:
                                                         entry_date = datetime.strptime(date_raw, '%Y/%m/%d').strftime('%Y-%m-%d')
                                                     except ValueError:
                                                          errors.append(f"Row {row_index}: Unrecognized date string format '{date_raw}'. Using today's date.")
                                                          entry_date = str(date.today())
                                 except ValueError: # Catches errors from fromisoformat or initial strptime
                                      errors.append(f"Row {row_index}: Unrecognized date string format '{date_raw}'. Using today's date.")
                                      entry_date = str(date.today())

                             else: # Handles other potential data types that aren't None
                                  errors.append(f"Row {row_index}: Unrecognized date data type '{type(date_raw).__name__}'. Using today's date.")
                                  entry_date = str(date.today())


                    # Basic validation for Article (Card is no longer required by mapping)
                    if not article:
                         errors.append(f"Row {row_index}: Missing required Article data. Skipping row.")
                         continue # Skip this row if Article is missing

                    data = (article, card, color, size, qty, component, print_opt, entry_date)

                    c.execute("INSERT INTO entries (article, card, color, size, qty, component, print_opt, date) VALUES (?, ?, ?, ?, ?, ?, ?, ?)", data)
                    imported_count += 1
                except IndexError:
                     errors.append(f"Row {row_index}: Data column index out of range based on mapping. Skipping row.")
                except Exception as e:
                     errors.append(f"Row {row_index}: Error processing row - {e}. Skipping row.")


            conn.commit()
            conn.close()

            success_message = f"Successfully imported {imported_count} entries."
            if errors:
                 error_message = "\n".join(errors)
                 messagebox.showwarning("Import with Errors", f"{success_message}\n\nErrors encountered:\n{error_message}")
            else:
                messagebox.showinfo("Import Complete", success_message)

            update_dashboard() # Refresh the view after import

        except FileNotFoundError:
            pass
        except Exception as e:
            messagebox.showerror("Import Error", f"Failed to process Excel file during import: {e}")


    def import_excel():
        filepath = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if not filepath: return

        try:
            workbook = load_workbook(filepath)
            sheet = workbook.active

            # Assuming headers are in row 2 based on your image
            header_row = 2
            if sheet.max_row < header_row:
                 messagebox.showerror("Import Error", f"Excel file does not have enough rows to read headers from row {header_row}.")
                 return

            # Read headers from the specified header row
            excel_headers = [sheet.cell(row=header_row, column=col+1).value for col in range(sheet.max_column)]
            excel_headers = [h for h in excel_headers if h is not None] # Filter out None values

            if not excel_headers:
                 messagebox.showerror("Import Error", "Could not find any headers in the specified header row of the Excel file.")
                 return

            # Show mapping window with discovered headers
            # Pass the header_row so perform_import knows where data starts
            show_import_mapping_window(filepath, excel_headers)

        except FileNotFoundError:
            pass
        except Exception as e:
            messagebox.showerror("Import Error", f"Could not read Excel headers or file: {e}")


    # --- UI Element Creation ---

    # --- Entry Form Frame ---
    entry_form_frame = LabelFrame(root, text="Entry Form", font=("Arial", 12, "bold"), bg="#e0f7fa", padx=10, pady=10)
    entry_form_frame.pack(pady=10, padx=10, fill="x")

    form_fields_labels = ["Article:", "Card:", "Color:", "Size:", "Qty:", "Component:"]
    article_entry = Entry(entry_form_frame, width=30)
    card_entry = Entry(entry_form_frame, width=30)
    color_entry = Entry(entry_form_frame, width=30)
    size_entry = Entry(entry_form_frame, width=30)
    qty_entry = Entry(entry_form_frame, width=30)
    component_entry = Entry(entry_form_frame, width=30)

    form_entries_widgets = [article_entry, card_entry, color_entry, size_entry, qty_entry, component_entry]

    for i, text in enumerate(form_fields_labels):
        Label(entry_form_frame, text=text, bg="#e0f7fa").grid(row=i, column=0, sticky="w", padx=5, pady=2)
        form_entries_widgets[i].grid(row=i, column=1, sticky="ew", padx=5, pady=2)

    Label(entry_form_frame, text="Print?:", bg="#e0f7fa").grid(row=len(form_fields_labels), column=0, sticky="w", padx=5, pady=2)
    print_var = StringVar(value="Yes")
    OptionMenu(entry_form_frame, print_var, "Yes", "Yes", "No").grid(row=len(form_fields_labels), column=1, sticky="ew", padx=5, pady=2)

    entry_form_frame.columnconfigure(1, weight=1)
    form_buttons_frame = Frame(entry_form_frame, bg="#e0f7fa")
    form_buttons_frame.grid(row=len(form_fields_labels) + 1, column=0, columnspan=2, pady=10)
    Button(form_buttons_frame, text="Save Entry", command=save_entry, bg="#00796b", fg="white", width=15).pack(side=LEFT, padx=5)
    Button(form_buttons_frame, text="Upload Image", command=upload_image, bg="#0288d1", fg="white", width=15).pack(side=LEFT, padx=5)

    # --- Search Section Frame ---
    search_controls_frame = LabelFrame(root, text="Search Entries", font=("Arial", 12, "bold"), bg="#b2ebf2", padx=10, pady=10)
    search_controls_frame.pack(pady=10, padx=10, fill="x")

    Label(search_controls_frame, text="Article:", bg="#b2ebf2").grid(row=0, column=0, padx=5, pady=5, sticky="w")
    search_article_entry = Entry(search_controls_frame)
    search_article_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

    Label(search_controls_frame, text="Card:", bg="#b2ebf2").grid(row=0, column=2, padx=5, pady=5, sticky="w")
    search_card_entry = Entry(search_controls_frame)
    search_card_entry.grid(row=0, column=3, padx=5, pady=5, sticky="ew")

    Label(search_controls_frame, text="Print:", bg="#b2ebf2").grid(row=0, column=4, padx=5, pady=5, sticky="w")
    search_print_var = StringVar(value="All")
    OptionMenu(search_controls_frame, search_print_var, "All", "All", "Yes", "No").grid(row=0, column=5, padx=5, pady=5, sticky="ew")

    Label(search_controls_frame, text="Start Date (YYYY-MM-DD):", bg="#b2ebf2").grid(row=1, column=0, padx=5, pady=5, sticky="w")
    search_start_date_entry = Entry(search_controls_frame)
    search_start_date_entry.grid(row=1, column=1, padx=5, pady=5, sticky="ew")

    Label(search_controls_frame, text="End Date (YYYY-MM-DD):", bg="#b2ebf2").grid(row=1, column=2, padx=5, pady=5, sticky="w")
    search_end_date_entry = Entry(search_controls_frame)
    search_end_date_entry.grid(row=1, column=3, padx=5, pady=5, sticky="ew")

    search_button = Button(search_controls_frame, text="Search", command=search_entries, bg="#00796b", fg="white")
    search_button.grid(row=1, column=4, columnspan=2, padx=5, pady=5, sticky="ew")

    search_controls_frame.columnconfigure(1, weight=1)
    search_controls_frame.columnconfigure(3, weight=1)
    search_controls_frame.columnconfigure(5, weight=1)

    # --- Data Display (Treeview) Frame ---
    data_display_frame = Frame(root, bg="#e0f7fa")
    data_display_frame.pack(pady=10, padx=10, fill="both", expand=True)

    tree_scrollbar_y = Scrollbar(data_display_frame)
    tree_scrollbar_y.pack(side=RIGHT, fill=Y)
    tree_scrollbar_x = Scrollbar(data_display_frame, orient=HORIZONTAL)
    tree_scrollbar_x.pack(side=BOTTOM, fill=X)

    tree = Treeview(data_display_frame, yscrollcommand=tree_scrollbar_y.set, xscrollcommand=tree_scrollbar_x.set)
    tree.pack(fill="both", expand=True)

    tree_scrollbar_y.config(command=tree.yview)
    tree_scrollbar_x.config(command=tree.xview)

    tree["columns"] = ("ID", "Article", "Card", "Color", "Size", "Qty", "Component", "Print", "Date")
    tree.column("#0", width=0, stretch=NO)
    tree.heading("#0", text="")

    for col in tree["columns"]:
        tree.column(col, anchor=CENTER, width=100)
        tree.heading(col, text=col)

    tree.column("ID", width=50)
    tree.column("Article", width=150)
    tree.column("Card", width=100)
    tree.column("Color", width=80)
    tree.column("Size", width=60)
    tree.column("Qty", width=50)
    tree.column("Component", width=150)
    tree.column("Print", width=60)
    tree.column("Date", width=100)


    # --- Action Buttons Frame ---
    action_buttons_frame = Frame(root, bg="#e0f7fa")
    action_buttons_frame.pack(pady=10)

    Button(action_buttons_frame, text="Edit Selected", command=edit_entry, bg="#ffb300", fg="white", width=15).pack(side=LEFT, padx=5)
    Button(action_buttons_frame, text="Delete Selected", command=delete_entry, bg="#d32f2f", fg="white", width=15).pack(side=LEFT, padx=5)
    Button(action_buttons_frame, text="Export to Excel", command=export_excel, bg="#388e3c", fg="white", width=15).pack(side=LEFT, padx=5)
    Button(action_buttons_frame, text="Export to PDF", command=export_pdf, bg="#c2185b", fg="white", width=15).pack(side=LEFT, padx=5)
    Button(action_buttons_frame, text="Import from Excel", command=import_excel, bg="#00acc1", fg="white", width=15).pack(side=LEFT, padx=5)


    init_db()
    update_dashboard()

    root.mainloop()

if __name__ == "__main__":
    show_main_ui()